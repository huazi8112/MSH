"""
Ranking-frequency distribution analysis.

Revised version
---------------
- Adds VoteRank, CHBC, and SNIM to the default method list.
- Uses the same revised method order, colors, markers, and PDF/PNG export style.
- Removes black marker outlines in frequency figures and legends to reduce occlusion when points overlap.
- Uses deterministic score grouping by rounded scores.
- Exports one combined Excel workbook plus per-network PDF/PNG figures.

The plot shows, for each method, how many nodes fall into each tied score-rank group.
It is a diagnostic of ranking discriminability and complements the scalar monotonicity
index M(R).

Run
---
python exp_ranking_frequency_revised_baselines_deterministic.py
python exp_ranking_frequency_revised_baselines_deterministic.py --networks jazz,email,usair
python exp_ranking_frequency_revised_baselines_deterministic.py --include-variants
"""

import argparse
import os
import random
from collections import Counter
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D

from hosh_methods import get_node_scores, get_network_partition
from network_loader import download_and_load_graph, get_network_list
from precompute_rankings import load_precomputed_rankings


MASTER_SEED = 42

CORE_METHODS = [
    "HOSH", "VoteRank", "SNIM", "CHBC",
    "ISH", "DC", "BC", "CC", "K-Shell", "SH", "CI", "SNC",
]

VARIANT_METHODS = [
    "HOSH-NO", "HOSH-NE", "HOSH-E", "HOSH-C",
    "HOSH-Lin", "HOSH-Sqrt", "HOSH-SumNorm",
]

COLORS = {
    "HOSH": "#D63230",
    "VoteRank": "#2CA02C",
    "SNIM": "#7F7F7F",
    "CHBC": "#5D3FD3",
    "ISH": "#F08C3D",
    "DC": "#E5B25D",
    "BC": "#4FA3D1",
    "CC": "#4364B8",
    "K-Shell": "#A855A8",
    "SH": "#E2739F",
    "CI": "#8D6E63",
    "SNC": "#4DB6AC",
    "HOSH-NO": "#A50F15",
    "HOSH-NE": "#FB6A4A",
    "HOSH-E": "#FC9272",
    "HOSH-C": "#FCAE91",
    "HOSH-Lin": "#756BB1",
    "HOSH-Sqrt": "#9E9AC8",
    "HOSH-SumNorm": "#BCBDDC",
}

MARKERS = {
    "HOSH": "o",
    "VoteRank": "o",
    "SNIM": "p",
    "CHBC": "*",
    "ISH": "s",
    "DC": "^",
    "BC": "D",
    "CC": "X",
    "K-Shell": "P",
    "SH": "v",
    "CI": "h",
    "SNC": "H",
    "HOSH-NO": "<",
    "HOSH-NE": ">",
    "HOSH-E": "8",
    "HOSH-C": "d",
    "HOSH-Lin": "1",
    "HOSH-Sqrt": "2",
    "HOSH-SumNorm": "3",
}

plt.rcParams.update({
    "font.family": "Times New Roman",
    "pdf.fonttype": 42,
    "ps.fonttype": 42,
    "font.size": 10,
    "axes.labelsize": 11,
    "axes.titlesize": 11,
    "xtick.labelsize": 9,
    "ytick.labelsize": 9,
    "legend.fontsize": 8.5,
    "figure.dpi": 600,
    "savefig.dpi": 600,
    "savefig.bbox": "tight",
    "savefig.pad_inches": 0.05,
    "axes.linewidth": 0.8,
    "lines.linewidth": 1.5,
    "lines.markersize": 4.5,
    "axes.grid": False,
    "axes.spines.top": True,
    "axes.spines.right": True,
    "xtick.major.width": 0.8,
    "ytick.major.width": 0.8,
    "xtick.major.size": 3.5,
    "ytick.major.size": 3.5,
})


# -----------------------------
# Utilities
# -----------------------------
def set_master_seed(seed: int = MASTER_SEED) -> None:
    random.seed(seed)
    np.random.seed(seed)


def parse_networks(network_arg: str) -> List[str]:
    if network_arg.strip():
        return [x.strip() for x in network_arg.split(",") if x.strip()]
    return get_network_list()


def parse_methods(method_arg: str, include_variants: bool) -> List[str]:
    if method_arg.strip():
        return [x.strip() for x in method_arg.split(",") if x.strip()]
    methods = list(CORE_METHODS)
    if include_variants:
        methods.extend(VARIANT_METHODS)
    return methods


def format_network_name(name: str) -> str:
    """Display network names consistently in figure titles."""
    s = str(name).strip()
    return s[:1].upper() + s[1:].lower() if s else s


def finite_float(x) -> Optional[float]:
    try:
        val = float(x)
        if np.isfinite(val):
            return val
    except Exception:
        pass
    return None


def load_or_compute_scores(method: str, graph, network_name: str,
                           partition=None, comm_size_map=None) -> Dict:
    precomputed = load_precomputed_rankings(network_name) if network_name else None
    if precomputed and method in precomputed and precomputed[method] is not None:
        return precomputed[method]

    if method == "HOSH-Sqrt" and precomputed and "HOSH-BoxCox" in precomputed:
        if precomputed["HOSH-BoxCox"] is not None:
            return precomputed["HOSH-BoxCox"]

    return get_node_scores(method, graph, partition=partition, comm_size_map=comm_size_map)


# -----------------------------
# Ranking frequency
# -----------------------------
def ranking_frequency_from_scores(scores: Dict, round_decimals: int = 8) -> Tuple[List[int], List[int], pd.DataFrame]:
    """
    Compute ranking-frequency distribution using competition ranks.
    Nodes with equal rounded score share the same rank.
    """
    score_counts = Counter()
    for _, score in scores.items():
        val = finite_float(score)
        if val is None:
            continue
        score_counts[round(val, round_decimals)] += 1

    sorted_scores = sorted(score_counts.keys(), reverse=True)
    records = []
    current_rank = 1
    for score in sorted_scores:
        freq = int(score_counts[score])
        records.append({
            "Ranking": current_rank,
            "Rounded_Score": score,
            "Frequency": freq,
        })
        current_rank += freq

    df = pd.DataFrame(records)
    return df["Ranking"].tolist(), df["Frequency"].tolist(), df


def analyze_network(network_name: str, methods: List[str], round_decimals: int):
    graph = download_and_load_graph(network_name, verbose=False)
    if graph is None or graph.number_of_nodes() == 0:
        print(f"  [Skip] {network_name}: failed to load or empty graph")
        return None, []

    print(f"  Nodes={graph.number_of_nodes()}, Edges={graph.number_of_edges()}")

    partition = None
    comm_size_map = None
    if "CHBC" in methods:
        try:
            partition, comm_size_map = get_network_partition(graph, seed=MASTER_SEED)
        except Exception as exc:
            print(f"  [Warning] CHBC partition failed: {exc}")

    result_map = {}
    long_rows = []

    for method in methods:
        print(f"    Method: {method}")
        try:
            scores = load_or_compute_scores(
                method, graph, network_name,
                partition=partition, comm_size_map=comm_size_map,
            )
            rankings, frequencies, df_freq = ranking_frequency_from_scores(scores, round_decimals=round_decimals)
            result_map[method] = {
                "rankings": rankings,
                "frequencies": frequencies,
                "df": df_freq,
            }
            for _, row in df_freq.iterrows():
                long_rows.append({
                    "Network": network_name,
                    "N": graph.number_of_nodes(),
                    "E": graph.number_of_edges(),
                    "Method": method,
                    "Ranking": int(row["Ranking"]),
                    "Rounded_Score": row["Rounded_Score"],
                    "Frequency": int(row["Frequency"]),
                })
        except Exception as exc:
            print(f"      [Failed] {method}: {exc}")

    return graph, result_map, long_rows


# -----------------------------
# Plotting and export
# -----------------------------
def export_standalone_legend(methods: List[str], output_dir: str) -> None:
    fig, ax = plt.subplots(figsize=(8.8, 1.1))
    ax.axis("off")
    handles = []
    for method in methods:
        handles.append(Line2D(
            [0], [0],
            marker=MARKERS.get(method, "o"),
            color="none",
            markerfacecolor=COLORS.get(method, "#000000"),
            markeredgecolor="none",
            markeredgewidth=0.0,
            markersize=6,
            label=method,
        ))
    leg = ax.legend(
        handles=handles,
        loc="center",
        ncol=6,
        frameon=True,
        fancybox=False,
        edgecolor="black",
        framealpha=1.0,
        columnspacing=0.8,
        handletextpad=0.35,
        borderpad=0.35,
    )
    leg.get_frame().set_linewidth(0.8)
    os.makedirs(output_dir, exist_ok=True)
    plt.savefig(os.path.join(output_dir, "RankingFreq_Legend_Standalone.pdf"), format="pdf")
    plt.savefig(os.path.join(output_dir, "RankingFreq_Legend_Standalone.png"), dpi=600, bbox_inches="tight", facecolor="white")
    plt.close()


def plot_ranking_frequency(network_name: str, graph, result_map: Dict, methods: List[str], output_dir: str,
                           embed_legend: bool = False) -> None:
    fig, ax = plt.subplots(figsize=(3.5, 2.8))

    for method in methods:
        if method not in result_map:
            continue
        rankings = result_map[method]["rankings"]
        frequencies = result_map[method]["frequencies"]
        if not rankings:
            continue

        if method == "HOSH":
            s, alpha, z = 42, 0.92, 10
        else:
            s, alpha, z = 25, 0.68, 5

        ax.scatter(
            rankings,
            frequencies,
            label=method,
            color=COLORS.get(method, "#000000"),
            marker=MARKERS.get(method, "o"),
            s=s,
            alpha=alpha,
            edgecolors="none",
            linewidths=0,
            zorder=z,
        )

    ax.set_title(format_network_name(network_name), fontsize=12, fontweight="bold", pad=8)
    ax.set_xlabel("Ranking")
    ax.set_ylabel("Frequency")

    all_rankings = [r for method in methods if method in result_map for r in result_map[method]["rankings"]]
    all_frequencies = [f for method in methods if method in result_map for f in result_map[method]["frequencies"]]
    if all_rankings:
        x_min, x_max = min(all_rankings), max(all_rankings)
        x_margin = max(1, (x_max - x_min) * 0.03)
        ax.set_xlim(x_min - x_margin, x_max + x_margin)
    if all_frequencies:
        y_max = max(all_frequencies)
        ax.set_ylim(0, y_max * 1.15 if y_max > 0 else 1)

    if embed_legend:
        ax.legend(
            loc="upper right",
            frameon=True,
            fancybox=False,
            shadow=False,
            edgecolor="black",
            framealpha=0.90,
            ncol=1,
            fontsize=6.8,
            borderpad=0.25,
            handletextpad=0.25,
            labelspacing=0.18,
        )

    for spine in ["left", "right", "top", "bottom"]:
        ax.spines[spine].set_linewidth(0.8)
        ax.spines[spine].set_color("#000000")
    ax.tick_params(direction="out", which="major", length=3.0, width=0.7)

    plt.tight_layout(pad=0.2)
    pdf_path = os.path.join(output_dir, f"RankingFreq_{network_name}.pdf")
    png_path = os.path.join(output_dir, f"RankingFreq_{network_name}.png")
    plt.savefig(pdf_path, format="pdf")
    plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white", edgecolor="none")
    plt.close()
    print(f"  [Output] {pdf_path}")


def export_results(long_rows: List[dict], output_dir: str, round_decimals: int, methods: List[str]) -> None:
    if not long_rows:
        return
    df_long = pd.DataFrame(long_rows)

    summary_rows = []
    for (network, method), sub in df_long.groupby(["Network", "Method"], sort=False):
        summary_rows.append({
            "Network": network,
            "Method": method,
            "N": int(sub["N"].iloc[0]),
            "E": int(sub["E"].iloc[0]),
            "Distinct_Rank_Groups": int(len(sub)),
            "Max_Frequency": int(sub["Frequency"].max()),
            "Mean_Frequency": float(sub["Frequency"].mean()),
            "Tied_Rank_Groups": int((sub["Frequency"] > 1).sum()),
            "Tied_Nodes": int(sub.loc[sub["Frequency"] > 1, "Frequency"].sum()),
        })
    df_summary = pd.DataFrame(summary_rows)

    notes = pd.DataFrame([
        {"Item": "Purpose", "Value": "Ranking-frequency distribution for score-rank groups."},
        {"Item": "Revised baselines", "Value": "VoteRank, SNIM, and CHBC are included by default."},
        {"Item": "Tie grouping", "Value": f"Scores are rounded to {round_decimals} decimals before grouping."},
        {"Item": "Ranking convention", "Value": "Competition ranking: nodes with equal rounded scores share the same rank; the next rank skips by group size."},
        {"Item": "Figures", "Value": "Per-network PDF/PNG figures plus a standalone legend; marker outlines are disabled to reduce overlap artifacts."},
    ])

    xlsx_path = os.path.join(output_dir, "Ranking_Frequency_Revised.xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_long.to_excel(writer, sheet_name="Long_Format", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        notes.to_excel(writer, sheet_name="Notes", index=False)

        # Optional network sheets for quick inspection.
        for network, sub_net in df_long.groupby("Network", sort=False):
            sheet_name = str(network)[:31]
            sub_net.to_excel(writer, sheet_name=sheet_name, index=False)

        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for ws in writer.book.worksheets:
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(cell.value, float):
                        cell.number_format = "0.000000"
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 16
            ws.freeze_panes = "A2"

    csv_path = os.path.join(output_dir, "Ranking_Frequency_Revised_Long.csv")
    df_long.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"[Output] Excel saved: {xlsx_path}")
    print(f"[Output] CSV saved: {csv_path}")


# -----------------------------
# Main
# -----------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Revised ranking-frequency distribution analysis.")
    parser.add_argument("--networks", type=str, default="",
                        help="Comma-separated network names. Default: get_network_list().")
    parser.add_argument("--methods", type=str, default="",
                        help="Comma-separated methods. Default: revised core methods.")
    parser.add_argument("--include-variants", action="store_true",
                        help="Also include HOSH ablation variants.")
    parser.add_argument("--round-decimals", type=int, default=8)
    parser.add_argument("--output-dir", type=str, default="results/exp_ranking_frequency_revised")
    parser.add_argument("--embed-legend", action="store_true",
                        help="Embed a small legend in each figure. Default: standalone legend only.")
    parser.add_argument("--master-seed", type=int, default=MASTER_SEED)
    args = parser.parse_args()

    set_master_seed(args.master_seed)
    networks = parse_networks(args.networks)
    methods = parse_methods(args.methods, args.include_variants)
    os.makedirs(args.output_dir, exist_ok=True)
    export_standalone_legend(methods, args.output_dir)

    print("=" * 72)
    print("Revised ranking-frequency distribution analysis")
    print(f"Networks: {networks}")
    print(f"Methods: {methods}")
    print("=" * 72)

    all_long_rows = []
    for idx, network in enumerate(networks, 1):
        print(f"\n[{idx}/{len(networks)}] Network: {network}")
        graph, result_map, long_rows = analyze_network(network, methods, args.round_decimals)
        if graph is None:
            continue
        plot_ranking_frequency(network, graph, result_map, methods, args.output_dir, embed_legend=args.embed_legend)
        all_long_rows.extend(long_rows)
        ckpt_path = os.path.join(args.output_dir, "Ranking_Frequency_Revised_Checkpoint.xlsx")
        pd.DataFrame(all_long_rows).to_excel(ckpt_path, index=False)
        print(f"  [Checkpoint] {ckpt_path}")

    export_results(all_long_rows, args.output_dir, args.round_decimals, methods)

    print("=" * 72)
    print("Completed.")
    print(f"Results saved to: {args.output_dir}")
    print("=" * 72)


if __name__ == "__main__":
    main()
