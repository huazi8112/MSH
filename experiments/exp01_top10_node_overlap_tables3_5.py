#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Top-10 node-overlap table for MSH, baseline methods, and SIR empirical ranking.

Purpose
-------
This script creates transposed Excel tables similar to a centrality-overlap table: each
network has methods as rows and Top-10 ranks as columns for MSH, baseline methods,
and the empirical single-node SIR ranking. Only nodes that appear in both MSH Top-10 and at least one other method
are highlighted with the same fill color. Overlaps among baseline methods
without MSH are intentionally not highlighted.

Default methods
---------------
MSH/HOSH, VoteRank, SNIM, CHBC, ISH, DC, BC, CC, K-Shell, SH, CI, SNC, and SIR.
HOSH is displayed as MSH in the output table.

SIR Top-10 source
-----------------
By default, the script first tries to read an existing SIR Top-10 CSV from
previous Top-10 overlap experiments. If no suitable CSV is found, it computes
single-node SIR influence from scratch and caches the result.

Default SIR protocol
--------------------
- gamma = 1.0
- beta = beta_multiplier * beta_threshold
- beta_threshold = <k> / (<k^2> - <k>)
- beta_multiplier = 2.5
- 50 blocks x 20 repeats per node
- discrete-time synchronous SIR

Outputs
-------
results/exp_top10_node_overlap_excel_table/
    Top10_Node_Overlap_Table.xlsx                  # updated after each network and at the end
    Top10_Node_Lists.csv                           # updated after each network and at the end
    MSH_Overlap_Summary.csv                        # updated after each network and at the end
    checkpoint_after_each_network/                 # per-network snapshots
    sir_cache/*.csv
"""
from __future__ import annotations

import argparse
import datetime as _dt
import hashlib
import math
import os
import random
import re
import sys
from collections import Counter
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import networkx as nx
import numpy as np
import pandas as pd
from tqdm import tqdm

from hosh_methods import get_node_scores
from network_loader import download_and_load_graph, get_network_list

try:
    from precompute_rankings import load_precomputed_rankings, get_standardized_ranked_nodes
except Exception:
    load_precomputed_rankings = None

    def get_standardized_ranked_nodes(scores: Dict, round_decimals: int = 8):
        return sorted(scores.keys(), key=lambda n: (-round(float(scores[n]), round_decimals), str(n)))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover
    raise ImportError("This script requires openpyxl for Excel formatting.") from exc


DEFAULT_METHODS = [
    "HOSH", "VoteRank", "SNIM", "CHBC", "ISH", "DC", "BC", "CC", "K-Shell", "SH", "CI", "SNC"
]

DUPLICATE_FILLS = [
    "92D050",  # green
    "FFE699",  # yellow
    "8EA9DB",  # blue
    "9999FF",  # purple
    "E2F0D9",  # pale green
    "FCE4D6",  # peach
    "D9EAD3",  # light green
    "DDEBF7",  # light blue
    "E4DFEC",  # lavender
    "F4CCCC",  # light red
    "FFF2CC",  # pale yellow
    "D0E0E3",  # cyan gray
]

DISPLAY_LABELS = {
    "HOSH": "MSH",
    "VoteRank": "VoteRank",
    "SNIM": "SNIM",
    "CHBC": "CHBC",
    "ISH": "ISH",
    "DC": "DC",
    "BC": "BC",
    "CC": "CC",
    "K-Shell": "K-Shell",
    "SH": "SH",
    "CI": "CI",
    "SNC": "SNC",
    "SIR": "SIR",
}

MASTER_SEED = 42


# =============================================================================
# Utilities
# =============================================================================
def stable_int_hash(*items: object, modulus: int = 2**32 - 1) -> int:
    text = "||".join(str(x) for x in items)
    digest = hashlib.blake2b(text.encode("utf-8"), digest_size=8).hexdigest()
    return int(digest, 16) % modulus


def parse_csv_list(text: str) -> List[str]:
    return [x.strip() for x in str(text).split(",") if x.strip()]


def safe_slug(text: object) -> str:
    s = str(text)
    s = re.sub(r"[^A-Za-z0-9_.-]+", "_", s)
    return s.strip("_") or "x"


def format_network_name(name: str) -> str:
    s = str(name)
    return s[:1].upper() + s[1:].lower() if s else s


def node_to_text(node) -> str:
    if pd.isna(node):
        return ""
    # Keep integer-like labels clean when values come from CSV/Excel.
    try:
        f = float(node)
        if f.is_integer() and str(node).strip().replace(".", "", 1).isdigit():
            return str(int(f))
    except Exception:
        pass
    return str(node)


def safe_sort_key(node) -> Tuple[str, str]:
    try:
        return ("0", f"{int(node):020d}")
    except Exception:
        return ("1", str(node))


def method_label(method: str) -> str:
    return DISPLAY_LABELS.get(method, method)


def standard_rank(scores: Dict) -> List:
    try:
        return list(get_standardized_ranked_nodes(scores, round_decimals=8))
    except Exception:
        return sorted(scores.keys(), key=lambda n: (-round(float(scores[n]), 8), safe_sort_key(n)))


# =============================================================================
# SIR ground truth
# =============================================================================
def beta_threshold_degree_moment(graph: nx.Graph) -> float:
    degrees = np.asarray([d for _, d in graph.degree()], dtype=float)
    if degrees.size == 0:
        return 0.0
    k1 = float(np.mean(degrees))
    k2 = float(np.mean(degrees ** 2))
    denom = k2 - k1
    if denom <= 0:
        return 1.0
    return k1 / denom


def resolve_beta(graph: nx.Graph, fixed_beta: Optional[float], beta_multiplier: float) -> Tuple[float, float, str]:
    beta_th = beta_threshold_degree_moment(graph)
    if fixed_beta is not None:
        beta = float(fixed_beta)
        mode = f"fixed beta={beta:.6f}"
    else:
        beta = float(beta_multiplier * beta_th)
        mode = f"beta={beta_multiplier:g}*beta_th"
    beta = max(0.0, min(1.0, beta))
    return beta, beta_th, mode


def run_sir_final(graph: nx.Graph, seeds: Iterable, beta: float, gamma: float,
                  rng_seed: int, max_steps: int = 1000) -> int:
    rng = random.Random(rng_seed)
    infected = set(n for n in seeds if graph.has_node(n))
    recovered = set()
    if not infected:
        return 0

    for _ in range(max_steps):
        if not infected:
            break
        new_infected = set()
        new_recovered = set()
        for u in list(infected):
            for v in graph.neighbors(u):
                if v not in infected and v not in recovered:
                    if rng.random() < beta:
                        new_infected.add(v)
            if rng.random() < gamma:
                new_recovered.add(u)
        infected.update(new_infected)
        infected.difference_update(new_recovered)
        recovered.update(new_recovered)
    return len(infected) + len(recovered)


def single_node_seed(network_name: str, beta: float, gamma: float, node,
                     block_idx: int, repeat_idx: int, master_seed: int) -> int:
    return int((master_seed + stable_int_hash(
        "single-node-sir", network_name, f"{beta:.8f}", f"{gamma:.8f}", str(node), block_idx, repeat_idx
    )) % (2**32 - 1))


def sir_cache_path(output_dir: str, network_name: str, beta: float, gamma: float,
                   blocks: int, repeats: int, master_seed: int) -> str:
    cache_dir = os.path.join(output_dir, "sir_cache")
    os.makedirs(cache_dir, exist_ok=True)
    fname = (
        f"SIR_NodeInfluence_{safe_slug(network_name)}_"
        f"beta{beta:.8f}_gamma{gamma:.8f}_b{blocks}_r{repeats}_seed{master_seed}.csv"
    ).replace(".", "p")
    return os.path.join(cache_dir, fname)


def compute_or_load_sir_ranking(graph: nx.Graph, network_name: str, beta: float, gamma: float,
                                blocks: int, repeats: int, master_seed: int,
                                max_steps: int, output_dir: str, use_cache: bool) -> List:
    path = sir_cache_path(output_dir, network_name, beta, gamma, blocks, repeats, master_seed)
    if use_cache and os.path.exists(path):
        df = pd.read_csv(path)
        return list(df.sort_values(["SIR_Final_Mean(%)", "Node_Text"], ascending=[False, True])["Node"])

    print(f"    Computing single-node SIR ranking for {graph.number_of_nodes()} nodes...")
    n = graph.number_of_nodes()
    records = []
    for node in tqdm(list(graph.nodes()), desc=f"    SIR nodes [{network_name}]", leave=False):
        block_values = []
        for b in range(blocks):
            total = 0
            for r in range(repeats):
                rng_seed = single_node_seed(network_name, beta, gamma, node, b, r, master_seed)
                total += run_sir_final(graph, [node], beta=beta, gamma=gamma, rng_seed=rng_seed, max_steps=max_steps)
            block_values.append((total / repeats) / n * 100.0)
        mean_val = float(np.mean(block_values)) if block_values else 0.0
        records.append({"Network": format_network_name(network_name), "Node": node, "Node_Text": node_to_text(node), "SIR_Final_Mean(%)": mean_val})

    df = pd.DataFrame(records)
    df = df.sort_values(["SIR_Final_Mean(%)", "Node_Text"], ascending=[False, True], kind="mergesort").reset_index(drop=True)
    df["SIR_Rank"] = np.arange(1, len(df) + 1)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return list(df["Node"])


def try_load_sir_topk_csv(network_name: str, top_k: int, sir_topk_csv: str) -> Optional[List[str]]:
    if not sir_topk_csv or not os.path.exists(sir_topk_csv):
        return None
    try:
        df = pd.read_csv(sir_topk_csv)
    except Exception:
        return None
    if df.empty or "Network" not in df.columns or "Node" not in df.columns:
        return None

    net_candidates = {str(network_name), format_network_name(network_name), str(network_name).lower(), format_network_name(network_name).lower()}
    sub = df[df["Network"].astype(str).str.lower().isin({x.lower() for x in net_candidates})].copy()
    if sub.empty:
        return None
    rank_col = "Rank" if "Rank" in sub.columns else None
    if rank_col:
        sub = sub.sort_values(rank_col, ascending=True)
    nodes = [node_to_text(x) for x in sub["Node"].tolist()[:top_k]]
    return nodes if len(nodes) >= min(top_k, len(sub)) else None


def default_sir_csv_candidates() -> List[str]:
    return [
        "results/exp_top10_overlap_with_sir/Top10_SIR_TopK.csv",
        "results/exp_msh_top10_overlap_with_methods_and_sir/Top10_SIR_TopK.csv",
        "results/exp_top10_overlap_with_sir/Top10_SIR_NodeInfluence.csv",
        "results/exp_msh_top10_overlap_with_methods_and_sir/Top10_SIR_NodeInfluence.csv",
    ]


def load_sir_topk_from_any_csv(network_name: str, top_k: int, explicit_path: str = "") -> Optional[List[str]]:
    candidates = [explicit_path] if explicit_path else []
    candidates.extend(default_sir_csv_candidates())
    for path in candidates:
        if not path or not os.path.exists(path):
            continue
        try:
            df = pd.read_csv(path)
        except Exception:
            continue
        if "Network" not in df.columns or "Node" not in df.columns:
            continue
        net_key = format_network_name(network_name).lower()
        sub = df[df["Network"].astype(str).str.lower() == net_key].copy()
        if sub.empty:
            continue
        if "Rank" in sub.columns:
            sub = sub.sort_values("Rank")
        elif "SIR_Rank" in sub.columns:
            sub = sub.sort_values("SIR_Rank")
        elif "SIR_Final_Mean(%)" in sub.columns:
            sub = sub.sort_values(["SIR_Final_Mean(%)", "Node"], ascending=[False, True])
        nodes = [node_to_text(x) for x in sub["Node"].tolist()[:top_k]]
        if nodes:
            print(f"    SIR Top-{top_k}: loaded from {path}")
            return nodes
    return None


# =============================================================================
# Ranking and overlap data
# =============================================================================
def load_or_compute_scores(method: str, graph: nx.Graph, network_name: str, use_precomputed: bool) -> Dict:
    if use_precomputed and load_precomputed_rankings is not None:
        rankings = load_precomputed_rankings(network_name)
        if rankings and method in rankings and rankings[method] is not None:
            return rankings[method]
    return get_node_scores(method, graph)


def collect_top10_nodes(methods: List[str], graph: nx.Graph, network_name: str, top_k: int,
                        args) -> Tuple[Dict[str, List[str]], Dict[str, object]]:
    top_nodes: Dict[str, List[str]] = {}
    for method in methods:
        print(f"    Ranking: {method}")
        scores = load_or_compute_scores(method, graph, network_name, args.use_precomputed)
        ranked = standard_rank(scores)
        top_nodes[method_label(method)] = [node_to_text(n) for n in ranked[:top_k]]

    # SIR empirical Top-10.
    sir_nodes = load_sir_topk_from_any_csv(network_name, top_k, args.sir_topk_csv)
    beta, beta_th, beta_mode = resolve_beta(graph, args.fixed_beta, args.beta_multiplier)
    if sir_nodes is None:
        sir_ranked = compute_or_load_sir_ranking(
            graph, network_name, beta=beta, gamma=args.gamma,
            blocks=args.blocks, repeats=args.repeats, master_seed=args.master_seed,
            max_steps=args.max_steps, output_dir=args.output_dir, use_cache=args.use_sir_cache,
        )
        sir_nodes = [node_to_text(n) for n in sir_ranked[:top_k]]
    top_nodes["SIR"] = sir_nodes[:top_k]

    meta = {"Beta": beta, "Beta_Threshold": beta_th, "Beta_Mode": beta_mode, "Gamma": args.gamma}
    return top_nodes, meta


def build_long_records(network_name: str, top_nodes: Dict[str, List[str]], top_k: int) -> List[Dict[str, object]]:
    records = []
    for method, nodes in top_nodes.items():
        for rank_idx in range(1, top_k + 1):
            node = nodes[rank_idx - 1] if rank_idx - 1 < len(nodes) else ""
            records.append({
                "Network": format_network_name(network_name),
                "Method": method,
                "Rank": rank_idx,
                "Node": node,
            })
    return records


def build_msh_overlap_records(network_name: str, top_nodes: Dict[str, List[str]], top_k: int) -> List[Dict[str, object]]:
    records = []
    msh_set = set(top_nodes.get("MSH", []))
    for method, nodes in top_nodes.items():
        if method == "MSH":
            continue
        target_set = set(nodes)
        inter = sorted(msh_set.intersection(target_set), key=safe_sort_key)
        union = msh_set.union(target_set)
        records.append({
            "Network": format_network_name(network_name),
            "Target": method,
            "TopK": top_k,
            "Overlap_Count": len(inter),
            "Overlap_Rate(%)": len(inter) / max(1, top_k) * 100.0,
            "Jaccard(%)": len(inter) / max(1, len(union)) * 100.0,
            "Shared_Nodes": ",".join(inter),
        })
    return records


# =============================================================================
# Excel formatting
# =============================================================================
def assign_duplicate_colors(top_nodes: Dict[str, List[str]]) -> Dict[str, str]:
    """Assign colors only to MSH-centered overlaps.

    A node is highlighted only when it appears in the MSH Top-K list and
    also appears in at least one other method, including SIR. Nodes repeated
    only among baseline methods are not highlighted.
    """
    msh_nodes = [node_to_text(n) for n in top_nodes.get("MSH", []) if node_to_text(n) != ""]
    msh_set = set(msh_nodes)

    other_nodes = []
    for method, nodes in top_nodes.items():
        if method == "MSH":
            continue
        other_nodes.extend([node_to_text(n) for n in nodes if node_to_text(n) != ""])

    other_counts = Counter(other_nodes)
    highlighted = [node for node in msh_nodes if node in other_counts]
    highlighted = sorted(set(highlighted), key=lambda n: (-other_counts[n], safe_sort_key(n)))
    return {node: DUPLICATE_FILLS[i % len(DUPLICATE_FILLS)] for i, node in enumerate(highlighted)}


def write_network_table(ws, start_row: int, table_idx: int, network_name: str,
                        methods: List[str], top_nodes: Dict[str, List[str]], top_k: int,
                        title_language: str = "en") -> int:
    """Write one transposed network table.

    Rows are methods and columns are ranks. Coloring is MSH-centered: a cell is
    highlighted only when its node appears in MSH Top-K and in at least one
    other method, including SIR.
    """
    labels = ["Method"] + [str(i) for i in range(1, top_k + 1)]
    n_cols = len(labels)
    end_col = get_column_letter(n_cols)

    if title_language.lower().startswith("zh"):
        title = f"表{table_idx}. {format_network_name(network_name)}网络中各方法确定的前十关键节点。"
    else:
        title = f"Table {table_idx}. Top-10 key nodes identified by different methods in the {format_network_name(network_name)} network."

    # Title row.
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=n_cols)
    cell = ws.cell(start_row, 1, title)
    cell.font = Font(name="Times New Roman", bold=True, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="808080")
    medium = Side(style="medium", color="808080")
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    header_row = start_row + 2
    for col_idx, label in enumerate(labels, 1):
        c = ws.cell(header_row, col_idx, label)
        c.font = Font(name="Times New Roman", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = header_fill
        c.border = Border(top=medium, bottom=thin, left=thin, right=thin)

    duplicate_color = assign_duplicate_colors(top_nodes)

    for row_offset, method in enumerate(methods, 1):
        row_idx = header_row + row_offset

        # Method name column.
        c_method = ws.cell(row_idx, 1, method)
        c_method.font = Font(name="Times New Roman", bold=True if method in {"HOSH", "MSH"} else False, size=10)
        c_method.alignment = Alignment(horizontal="center", vertical="center")
        c_method.border = Border(left=thin, right=thin)
        if method in {"HOSH", "MSH"}:
            c_method.fill = PatternFill("solid", fgColor="F2F2F2")

        nodes = top_nodes.get(method, [])
        for rank_idx in range(top_k):
            node = nodes[rank_idx] if rank_idx < len(nodes) else ""
            c = ws.cell(row_idx, rank_idx + 2, node)
            c.font = Font(name="Times New Roman", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=thin, right=thin)
            if node in duplicate_color:
                c.fill = PatternFill("solid", fgColor=duplicate_color[node])

    bottom_row = header_row + len(methods)
    for col_idx in range(1, n_cols + 1):
        c = ws.cell(bottom_row, col_idx)
        c.border = Border(bottom=medium, left=thin, right=thin)

    # Top and bottom horizontal separator lines around each table.
    for col_idx in range(1, n_cols + 1):
        ws.cell(start_row - 1, col_idx).border = Border(bottom=medium)
        ws.cell(bottom_row + 1, col_idx).border = Border(bottom=medium)

    return bottom_row + 3

def style_summary_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Times New Roman", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Times New Roman", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        width = max(10, min(34, max(len(str(ws.cell(row, col).value or "")) for row in range(1, min(ws.max_row, 50) + 1)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def save_excel(all_tables: Dict[str, Dict[str, List[str]]], methods: List[str], top_k: int,
               long_df: pd.DataFrame, overlap_df: pd.DataFrame, meta_df: pd.DataFrame,
               output_path: str, title_language: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top10_Overlap_Tables"
    ws.sheet_view.showGridLines = False

    row = 2
    for idx, (network_name, top_nodes) in enumerate(all_tables.items(), 1):
        row = write_network_table(ws, row, idx, network_name, methods, top_nodes, top_k, title_language=title_language)

    # Column widths for transposed visual table.
    ws.column_dimensions["A"].width = 16
    for col in range(2, top_k + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    # Raw Top-10 node list.
    ws_long = wb.create_sheet("Top10_Nodes")
    ws_long.append(list(long_df.columns))
    for row_values in long_df.itertuples(index=False):
        ws_long.append(list(row_values))
    style_summary_sheet(ws_long)

    # MSH overlap summary.
    ws_overlap = wb.create_sheet("MSH_Overlap_Summary")
    ws_overlap.append(list(overlap_df.columns))
    for row_values in overlap_df.itertuples(index=False):
        ws_overlap.append(list(row_values))
    style_summary_sheet(ws_overlap)

    # Protocol metadata.
    ws_meta = wb.create_sheet("Protocol_Metadata")
    ws_meta.append(list(meta_df.columns))
    for row_values in meta_df.itertuples(index=False):
        ws_meta.append(list(row_values))
    style_summary_sheet(ws_meta)

    wb.save(output_path)



# =============================================================================
# Incremental output helper
# =============================================================================
def write_incremental_outputs(all_tables: Dict[str, Dict[str, List[str]]],
                              display_methods: List[str], top_k: int,
                              long_records: List[Dict[str, object]],
                              overlap_records: List[Dict[str, object]],
                              meta_records: List[Dict[str, object]],
                              output_dir: str, title_language: str,
                              checkpoint_network: Optional[str] = None) -> Tuple[str, str, str]:
    """Write current accumulated results immediately.

    The main workbook and CSV files are overwritten after each completed network,
    so partial results are available even while later networks are still running.
    A network-specific checkpoint workbook is also saved for inspection.
    """
    os.makedirs(output_dir, exist_ok=True)
    long_df = pd.DataFrame(long_records)
    overlap_df = pd.DataFrame(overlap_records)
    meta_df = pd.DataFrame(meta_records)

    long_csv = os.path.join(output_dir, "Top10_Node_Lists.csv")
    overlap_csv = os.path.join(output_dir, "MSH_Overlap_Summary.csv")
    xlsx_path = os.path.join(output_dir, "Top10_Node_Overlap_Table.xlsx")

    long_df.to_csv(long_csv, index=False, encoding="utf-8-sig")
    overlap_df.to_csv(overlap_csv, index=False, encoding="utf-8-sig")
    save_excel(all_tables, display_methods, top_k, long_df, overlap_df, meta_df, xlsx_path, title_language)

    checkpoint_path = xlsx_path
    if checkpoint_network:
        checkpoint_dir = os.path.join(output_dir, "checkpoint_after_each_network")
        os.makedirs(checkpoint_dir, exist_ok=True)
        checkpoint_path = os.path.join(
            checkpoint_dir,
            f"Top10_Node_Overlap_Table_until_{safe_slug(format_network_name(checkpoint_network))}.xlsx",
        )
        save_excel(all_tables, display_methods, top_k, long_df, overlap_df, meta_df, checkpoint_path, title_language)

    return xlsx_path, long_csv, overlap_csv

# =============================================================================
# Main
# =============================================================================
def parse_args():
    parser = argparse.ArgumentParser(description="Create Excel Top-10 node overlap tables for MSH, baselines, and SIR.")
    parser.add_argument("--output-dir", default="results/exp_top10_node_overlap_excel_table")
    parser.add_argument("--networks", default="", help="Comma-separated networks. Empty means get_network_list().")
    parser.add_argument("--methods", default=",".join(DEFAULT_METHODS), help="Comma-separated ranking methods. HOSH is displayed as MSH.")
    parser.add_argument("--top-k", type=int, default=10)
    parser.add_argument("--sir-topk-csv", default="", help="Optional existing CSV containing columns Network, Rank, Node for SIR Top-K.")
    parser.add_argument("--gamma", type=float, default=1.0)
    parser.add_argument("--beta-multiplier", type=float, default=2.5)
    parser.add_argument("--fixed-beta", type=float, default=None)
    parser.add_argument("--blocks", type=int, default=50)
    parser.add_argument("--repeats", type=int, default=20)
    parser.add_argument("--max-steps", type=int, default=1000)
    parser.add_argument("--master-seed", type=int, default=42)
    parser.set_defaults(use_precomputed=True)
    parser.add_argument("--fresh-rankings", dest="use_precomputed", action="store_false")
    parser.set_defaults(use_sir_cache=True)
    parser.add_argument("--fresh-sir", dest="use_sir_cache", action="store_false")
    parser.add_argument("--title-language", choices=["en", "zh"], default="en")
    return parser.parse_args()


def main():
    args = parse_args()
    os.makedirs(args.output_dir, exist_ok=True)

    networks = parse_csv_list(args.networks) if args.networks else get_network_list()
    raw_methods = parse_csv_list(args.methods)
    if "HOSH" not in raw_methods:
        raw_methods = ["HOSH"] + raw_methods

    display_methods = [method_label(m) for m in raw_methods] + ["SIR"]
    # Keep unique display labels while preserving order.
    seen = set()
    display_methods = [m for m in display_methods if not (m in seen or seen.add(m))]

    print("=" * 80)
    print("Top-10 node-overlap Excel table")
    print("=" * 80)
    print(f"Networks: {networks}")
    print(f"Methods: {display_methods}")
    print(f"Top-K: {args.top_k}")
    print(f"SIR protocol: gamma={args.gamma}, beta_multiplier={args.beta_multiplier}, blocks={args.blocks}, repeats={args.repeats}")
    print("=" * 80)

    all_tables: Dict[str, Dict[str, List[str]]] = {}
    long_records: List[Dict[str, object]] = []
    overlap_records: List[Dict[str, object]] = []
    meta_records: List[Dict[str, object]] = []

    for idx, net in enumerate(networks, 1):
        print(f"\n[{idx}/{len(networks)}] Network: {net}")
        graph = download_and_load_graph(net, verbose=False)
        if graph is None or graph.number_of_nodes() == 0:
            print(f"  [Skip] {net}: empty or failed to load.")
            continue

        top_nodes, beta_meta = collect_top10_nodes(raw_methods, graph, net, args.top_k, args)
        all_tables[net] = top_nodes
        long_records.extend(build_long_records(net, top_nodes, args.top_k))
        overlap_records.extend(build_msh_overlap_records(net, top_nodes, args.top_k))
        meta_records.append({
            "Network": format_network_name(net),
            "N": graph.number_of_nodes(),
            "E": graph.number_of_edges(),
            "TopK": args.top_k,
            "Use_Precomputed_Rankings": args.use_precomputed,
            "SIR_Beta": beta_meta["Beta"],
            "SIR_Beta_Threshold": beta_meta["Beta_Threshold"],
            "SIR_Beta_Mode": beta_meta["Beta_Mode"],
            "SIR_Gamma": beta_meta["Gamma"],
            "SIR_Blocks": args.blocks,
            "SIR_Repeats": args.repeats,
            "Generated_At": _dt.datetime.now().isoformat(timespec="seconds"),
            "Python": sys.version.replace("\n", " "),
        })

        # Write current results immediately after each completed network.
        xlsx_path, long_csv, overlap_csv = write_incremental_outputs(
            all_tables=all_tables,
            display_methods=display_methods,
            top_k=args.top_k,
            long_records=long_records,
            overlap_records=overlap_records,
            meta_records=meta_records,
            output_dir=args.output_dir,
            title_language=args.title_language,
            checkpoint_network=net,
        )
        print(f"  [Saved] Current Excel: {xlsx_path}")
        print(f"  [Saved] Current CSV: {long_csv}")
        print(f"  [Saved] Checkpoint for {format_network_name(net)} saved under checkpoint_after_each_network/")

    if not all_tables:
        print("No valid network table was generated.")
        return

    xlsx_path, long_csv, overlap_csv = write_incremental_outputs(
        all_tables=all_tables,
        display_methods=display_methods,
        top_k=args.top_k,
        long_records=long_records,
        overlap_records=overlap_records,
        meta_records=meta_records,
        output_dir=args.output_dir,
        title_language=args.title_language,
        checkpoint_network=None,
    )

    print("\n" + "=" * 80)
    print("Done.")
    print(f"Final Excel saved: {xlsx_path}")
    print(f"Final CSV saved: {long_csv}")
    print(f"Overlap summary saved: {overlap_csv}")
    print(f"Per-network checkpoints saved to: {os.path.join(args.output_dir, 'checkpoint_after_each_network')}")
    print("=" * 80)


if __name__ == "__main__":
    main()
