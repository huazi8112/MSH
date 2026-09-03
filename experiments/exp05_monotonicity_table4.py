"""
Monotonicity analysis for node-ranking methods.

Manuscript-table version
------------------------
- Computes monotonicity directly from score ties, not from tie-broken rankings.
- No ID-based tie-breaking is used in the monotonicity calculation.
- Generates a compact manuscript table with methods as rows and networks as columns.
- Keeps detailed tie diagnostics in separate sheets for appendix/checking.

Metric
------
M(R) = (1 - sum_r n_r (n_r - 1) / (N (N - 1)))^2,
where n_r is the number of nodes sharing the same score value.
A larger value indicates fewer tied scores and stronger ranking resolution.

Run
---
python exp_monotonicity_revised_baselines_no_tiebreak_manuscript_table.py
python exp_monotonicity_revised_baselines_no_tiebreak_manuscript_table.py --networks jazz,email,usair
python exp_monotonicity_revised_baselines_no_tiebreak_manuscript_table.py --include-variants
python exp_monotonicity_revised_baselines_no_tiebreak_manuscript_table.py --no-rounding
"""

import argparse
import os
import random
from typing import Dict, List, Optional

import numpy as np
import pandas as pd

from hosh_methods import get_node_scores, get_network_partition
from network_loader import download_and_load_graph, get_network_list
from precompute_rankings import load_precomputed_rankings


MASTER_SEED = 42

CORE_METHODS = [
    "HOSH", "VoteRank", "SNIM", "CHBC",
    "ISH", "DC", "BC", "CC", "K-Shell", "SH", "CI", "SNC",
]

VARIANT_METHODS = [
    "HOSH-NO", "HOSH-NE", "HOSH-E", "HOSH-C",
    "HOSH-Lin", "HOSH-Sqrt", "HOSH-SumNorm",
]

METHOD_GROUP = {
    "HOSH": "Proposed",
    "VoteRank": "Diversified baseline",
    "SNIM": "Clique-based baseline",
    "CHBC": "Community-aware baseline",
    "ISH": "Structural-hole baseline",
    "DC": "Centrality baseline",
    "BC": "Centrality baseline",
    "CC": "Centrality baseline",
    "K-Shell": "Centrality baseline",
    "SH": "Structural-hole baseline",
    "CI": "Local/global influence baseline",
    "SNC": "Neighborhood baseline",
    "HOSH-NO": "Ablation variant",
    "HOSH-NE": "Ablation variant",
    "HOSH-E": "Ablation variant",
    "HOSH-C": "Ablation variant",
    "HOSH-Lin": "Ablation variant",
    "HOSH-Sqrt": "Ablation variant",
    "HOSH-SumNorm": "Ablation variant",
}

# Display names used in manuscript tables. Internal method names are unchanged.
DISPLAY_NAME = {
    "HOSH": "MSH",
    "HOSH-NO": "MSH-NO",
    "HOSH-NE": "MSH-NE",
    "HOSH-E": "MSH-E",
    "HOSH-C": "MSH-C",
    "HOSH-Lin": "MSH-Lin",
    "HOSH-Sqrt": "MSH-Sqrt",
    "HOSH-SumNorm": "MSH-SumNorm",
}

NETWORK_DISPLAY = {
    "lesmis": "Lesmis",
    "adjnoun": "Adjnoun",
    "adjoun": "Adjoun",
    "jazz": "Jazz",
    "usair": "USAir",
    "infect": "Infect",
    "email": "Email",
    "polblogs": "Polblogs",
    "hamster": "Hamster",
    "power": "Power",
}


def display_method(method: str) -> str:
    return DISPLAY_NAME.get(method, method)


def display_network(network: str) -> str:
    return NETWORK_DISPLAY.get(str(network).lower(), str(network).capitalize())


# -----------------------------
# Reproducibility utilities
# -----------------------------
def set_master_seed(seed: int = MASTER_SEED) -> None:
    random.seed(seed)
    np.random.seed(seed)


def parse_networks(network_arg: str) -> List[str]:
    if network_arg.strip():
        return [x.strip() for x in network_arg.split(",") if x.strip()]
    return get_network_list()


def parse_methods(method_arg: str, include_variants: bool) -> List[str]:
    if method_arg.strip():
        return [x.strip() for x in method_arg.split(",") if x.strip()]
    methods = list(CORE_METHODS)
    if include_variants:
        methods.extend(VARIANT_METHODS)
    return methods


def finite_float(x) -> Optional[float]:
    try:
        val = float(x)
        if np.isfinite(val):
            return val
    except Exception:
        pass
    return None


# -----------------------------
# Scores and monotonicity metric
# -----------------------------
def load_or_compute_scores(method: str, graph, network_name: str,
                           partition=None, comm_size_map=None) -> Dict:
    """Load precomputed scores when available; otherwise compute on the fly."""
    precomputed = load_precomputed_rankings(network_name) if network_name else None
    if precomputed and method in precomputed and precomputed[method] is not None:
        return precomputed[method]

    # Compatibility for older precomputed naming.
    if method == "HOSH-Sqrt" and precomputed and "HOSH-BoxCox" in precomputed:
        if precomputed["HOSH-BoxCox"] is not None:
            return precomputed["HOSH-BoxCox"]

    return get_node_scores(method, graph, partition=partition, comm_size_map=comm_size_map)


def score_group_counts(scores: Dict, round_decimals: Optional[int] = 10) -> List[int]:
    """
    Return the counts of nodes sharing the same score value.

    Important: this function does not apply any ID-based tie-breaking. Tied nodes
    remain tied. Optional rounding only avoids artificial floating-point noise.
    Set round_decimals=None to group by raw floating-point values.
    """
    grouped = {}
    for _, score in scores.items():
        val = finite_float(score)
        if val is None:
            continue
        key = val if round_decimals is None else round(val, round_decimals)
        grouped[key] = grouped.get(key, 0) + 1
    return list(grouped.values())


def monotonicity_mr(scores: Dict, n_nodes: int,
                    round_decimals: Optional[int] = 10) -> Dict[str, float]:
    """Compute M(R) from score groups without tie-breaking."""
    if n_nodes <= 1:
        return {
            "M(R)": 1.0,
            "Tie_Groups": 0,
            "Tied_Nodes": 0,
            "Max_Tie_Group_Size": 1,
            "Distinct_Score_Groups": n_nodes,
        }

    counts = score_group_counts(scores, round_decimals=round_decimals)
    if not counts:
        return {
            "M(R)": np.nan,
            "Tie_Groups": np.nan,
            "Tied_Nodes": np.nan,
            "Max_Tie_Group_Size": np.nan,
            "Distinct_Score_Groups": np.nan,
        }

    counts_arr = np.asarray(counts, dtype=float)
    same_rank_pairs = float(np.sum(counts_arr * (counts_arr - 1)))
    total_pairs = float(n_nodes * (n_nodes - 1))
    mr = (1.0 - same_rank_pairs / total_pairs) ** 2 if total_pairs > 0 else 1.0

    tie_groups = int(np.sum(counts_arr > 1))
    tied_nodes = int(np.sum(counts_arr[counts_arr > 1]))
    max_tie = int(np.max(counts_arr)) if len(counts_arr) else 0

    return {
        "M(R)": float(mr),
        "Tie_Groups": tie_groups,
        "Tied_Nodes": tied_nodes,
        "Max_Tie_Group_Size": max_tie,
        "Distinct_Score_Groups": int(len(counts_arr)),
    }


def analyze_network(network_name: str, methods: List[str],
                    round_decimals: Optional[int]) -> List[dict]:
    graph = download_and_load_graph(network_name, verbose=False)
    if graph is None or graph.number_of_nodes() == 0:
        print(f"  [Skip] {network_name}: failed to load or empty graph")
        return []

    n = graph.number_of_nodes()
    e = graph.number_of_edges()
    print(f"  Nodes={n}, Edges={e}")

    partition = None
    comm_size_map = None
    if "CHBC" in methods:
        try:
            partition, comm_size_map = get_network_partition(graph, seed=MASTER_SEED)
        except Exception as exc:
            print(f"  [Warning] CHBC partition failed: {exc}")

    rows = []
    for method in methods:
        print(f"    Method: {method}")
        try:
            scores = load_or_compute_scores(
                method, graph, network_name,
                partition=partition, comm_size_map=comm_size_map,
            )
            stats = monotonicity_mr(scores, n, round_decimals=round_decimals)
            rows.append({
                "Network": network_name,
                "Network_Display": display_network(network_name),
                "N": n,
                "E": e,
                "Method": method,
                "Method_Display": display_method(method),
                "Method_Group": METHOD_GROUP.get(method, "Other"),
                **stats,
            })
        except Exception as exc:
            print(f"      [Failed] {method}: {exc}")
            rows.append({
                "Network": network_name,
                "Network_Display": display_network(network_name),
                "N": n,
                "E": e,
                "Method": method,
                "Method_Display": display_method(method),
                "Method_Group": METHOD_GROUP.get(method, "Other"),
                "M(R)": np.nan,
                "Tie_Groups": np.nan,
                "Tied_Nodes": np.nan,
                "Max_Tie_Group_Size": np.nan,
                "Distinct_Score_Groups": np.nan,
            })
    return rows


# -----------------------------
# Tables
# -----------------------------
def build_network_by_method_table(df_long: pd.DataFrame, methods: List[str]) -> pd.DataFrame:
    """Detailed wide table: one row per network, one column per method."""
    if df_long.empty:
        return pd.DataFrame()
    base = df_long[["Network", "Network_Display", "N", "E"]].drop_duplicates().reset_index(drop=True)
    wide = df_long.pivot(index="Network", columns="Method", values="M(R)").reset_index()
    table = base.merge(wide, on="Network", how="left")
    cols = ["Network", "Network_Display", "N", "E"] + [m for m in methods if m in table.columns]
    return table[cols]


def build_manuscript_table(df_long: pd.DataFrame, methods: List[str],
                           networks: List[str]) -> pd.DataFrame:
    """
    Compact manuscript table: methods as rows, networks as columns, plus Mean.
    This layout is more suitable when many methods are compared.
    """
    if df_long.empty:
        return pd.DataFrame()

    network_cols = [display_network(n) for n in networks]
    pivot = df_long.pivot(index="Method", columns="Network_Display", values="M(R)")

    rows = []
    for method in methods:
        if method not in pivot.index:
            continue
        row = {"Method": display_method(method)}
        vals = []
        for col in network_cols:
            val = pivot.loc[method, col] if col in pivot.columns else np.nan
            row[col] = val
            if pd.notna(val):
                vals.append(float(val))
        row["Mean"] = float(np.mean(vals)) if vals else np.nan
        rows.append(row)
    return pd.DataFrame(rows, columns=["Method"] + network_cols + ["Mean"])


def build_method_summary(df_long: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for method, sub in df_long.groupby("Method", sort=False):
        vals = sub["M(R)"].to_numpy(dtype=float)
        rows.append({
            "Method": method,
            "Method_Display": display_method(method),
            "Method_Group": sub["Method_Group"].iloc[0],
            "Mean_M(R)": float(np.nanmean(vals)),
            "Std_M(R)": float(np.nanstd(vals, ddof=1)) if np.sum(~np.isnan(vals)) > 1 else np.nan,
            "Mean_Tie_Groups": float(np.nanmean(sub["Tie_Groups"].to_numpy(dtype=float))),
            "Mean_Tied_Nodes": float(np.nanmean(sub["Tied_Nodes"].to_numpy(dtype=float))),
            "Mean_Max_Tie_Group_Size": float(np.nanmean(sub["Max_Tie_Group_Size"].to_numpy(dtype=float))),
            "Mean_Distinct_Score_Groups": float(np.nanmean(sub["Distinct_Score_Groups"].to_numpy(dtype=float))),
        })
    return pd.DataFrame(rows)


# -----------------------------
# Export
# -----------------------------
def export_results(df_long: pd.DataFrame, methods: List[str], networks: List[str],
                   output_dir: str, round_decimals: Optional[int], master_seed: int) -> None:
    os.makedirs(output_dir, exist_ok=True)

    detailed_wide = build_network_by_method_table(df_long, methods)
    manuscript_all = build_manuscript_table(df_long, methods, networks)
    manuscript_core = build_manuscript_table(df_long, [m for m in CORE_METHODS if m in methods], networks)
    summary = build_method_summary(df_long)

    grouping_note = "Raw score values are grouped directly; no rounding is applied." if round_decimals is None else (
        f"Score values are rounded to {round_decimals} decimals only to suppress floating-point noise. No ID-based tie-breaking is applied."
    )
    notes = pd.DataFrame([
        {"Item": "Metric", "Value": "M(R) = (1 - sum_r n_r(n_r-1)/(N(N-1)))^2"},
        {"Item": "Score grouping", "Value": grouping_note},
        {"Item": "Interpretation", "Value": "Higher M(R) indicates fewer tied score groups and stronger ranking resolution."},
        {"Item": "Main-text table", "Value": "Use Manuscript_Core_Table. Method rows and network columns reduce table width."},
        {"Item": "Variants", "Value": "If variants are included, use Manuscript_All_Methods_Table or Method_Summary in the appendix."},
        {"Item": "Randomness", "Value": f"No SIR simulation is used; master seed={master_seed} is set for deterministic fallback computations."},
    ])

    xlsx_path = os.path.join(output_dir, "Monotonicity_Results_NoTieBreak_Manuscript.xlsx")
    csv_long_path = os.path.join(output_dir, "Monotonicity_Results_NoTieBreak_Long.csv")
    csv_table_path = os.path.join(output_dir, "Monotonicity_Manuscript_Core_Table.csv")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        manuscript_core.to_excel(writer, sheet_name="Manuscript_Core_Table", index=False)
        manuscript_all.to_excel(writer, sheet_name="Manuscript_All_Methods", index=False)
        detailed_wide.to_excel(writer, sheet_name="Network_by_Method", index=False)
        summary.to_excel(writer, sheet_name="Method_Summary", index=False)
        df_long.to_excel(writer, sheet_name="Long_Format", index=False)
        notes.to_excel(writer, sheet_name="Notes", index=False)

        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        thin = Side(style="thin", color="BFBFBF")
        medium = Side(style="medium", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="F2F2F2")
        proposed_fill = PatternFill("solid", fgColor="FFF2CC")

        for ws in writer.book.worksheets:
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "B2" if ws.title.startswith("Manuscript") else "A2"

            # Header style
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=medium, bottom=thin)

            # Body style
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                is_proposed = str(row[0].value) in {"MSH", "HOSH"}
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(cell.value, float):
                        cell.number_format = "0.0000"
                    if is_proposed and ws.title.startswith("Manuscript"):
                        cell.fill = proposed_fill
                        cell.font = Font(bold=True)

            # Bottom border for final row
            if ws.max_row >= 2:
                for cell in ws[ws.max_row]:
                    cell.border = Border(bottom=medium, left=thin, right=thin, top=thin)

            # Widths
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                width = 12
                if col_idx == 1:
                    width = 16
                if ws.title == "Long_Format":
                    width = 18
                ws.column_dimensions[letter].width = width

    df_long.to_csv(csv_long_path, index=False, encoding="utf-8-sig")
    manuscript_core.to_csv(csv_table_path, index=False, encoding="utf-8-sig")

    print(f"[Output] Excel saved: {xlsx_path}")
    print(f"[Output] Manuscript CSV saved: {csv_table_path}")
    print(f"[Output] Long CSV saved: {csv_long_path}")


# -----------------------------
# Main
# -----------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Monotonicity M(R) analysis without tie-breaking.")
    parser.add_argument("--networks", type=str, default="",
                        help="Comma-separated network names. Default: get_network_list().")
    parser.add_argument("--methods", type=str, default="",
                        help="Comma-separated method list. Default: revised core methods.")
    parser.add_argument("--include-variants", action="store_true",
                        help="Also include MSH ablation variants. Recommended for appendix only.")
    parser.add_argument("--round-decimals", type=int, default=10,
                        help="Rounding precision for score grouping. Used only to suppress floating-point noise.")
    parser.add_argument("--no-rounding", action="store_true",
                        help="Group by raw score values with no rounding.")
    parser.add_argument("--output-dir", type=str, default="results/exp_monotonicity_no_tiebreak")
    parser.add_argument("--master-seed", type=int, default=MASTER_SEED)
    args = parser.parse_args()

    set_master_seed(args.master_seed)
    networks = parse_networks(args.networks)
    methods = parse_methods(args.methods, args.include_variants)
    round_decimals = None if args.no_rounding else args.round_decimals

    print("=" * 72)
    print("Monotonicity analysis without ID-based tie-breaking")
    print(f"Networks: {networks}")
    print(f"Methods: {methods}")
    print(f"Round decimals: {round_decimals}")
    print("=" * 72)

    all_rows = []
    os.makedirs(args.output_dir, exist_ok=True)

    for idx, net in enumerate(networks, 1):
        print(f"\n[{idx}/{len(networks)}] Network: {net}")
        rows = analyze_network(net, methods, round_decimals=round_decimals)
        if rows:
            all_rows.extend(rows)
            df_ckpt = pd.DataFrame(all_rows)
            ckpt_path = os.path.join(args.output_dir, "Monotonicity_NoTieBreak_Checkpoint.xlsx")
            df_ckpt.to_excel(ckpt_path, index=False)
            print(f"  [Checkpoint] {ckpt_path}")

    if not all_rows:
        print("No valid results generated.")
        return

    df_long = pd.DataFrame(all_rows)
    export_results(df_long, methods, networks, args.output_dir, round_decimals, args.master_seed)

    print("=" * 72)
    print("Completed.")
    print(f"Results saved to: {args.output_dir}")
    print("=" * 72)


if __name__ == "__main__":
    main()
