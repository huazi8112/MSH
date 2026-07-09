"""
Ablation experiment: average improvement table with statistical significance for p = 1%, 5%, 10%.

Purpose
-------
This script evaluates the full HOSH/MSH model and its ablation variants under the
SIR final infection scale, then reports the improvement percentage of the full model over
each variant. The main manuscript table reports the average improvement over networks
for p = 1%, 5%, and 10%, with Wilcoxon significance markers.

Network | p | HOSH-NO | HOSH-NE | HOSH-E | HOSH-C | HOSH-Lin | HOSH-Sqrt | HOSH-SumNorm

Improvement percentage is defined as:
    eta = (F_HOSH - F_variant) / F_variant * 100%

Positive eta means the full HOSH model outperforms the variant.
Negative eta means the variant has a larger final infection scale.

Output
------
results/exp_ablation_improvement_sig/ablation_average_improvement_significance.xlsx

Notes
-----
- Uses master seed = 42.
- Uses common random numbers across HOSH and all variants under the same
  network, seed ratio, simulation block, and repeat index.
- Common random numbers are used across the full model and all variants.
- Network-level paired Wilcoxon tests are used for the manuscript summary table.
- Block-level paired Wilcoxon tests are also saved for detailed checking.
"""

import argparse
import os
import random
import hashlib
from typing import Dict, Iterable, List, Tuple

import numpy as np
import pandas as pd
from scipy.stats import t, wilcoxon
from tqdm import tqdm

from hosh_methods import get_node_scores
from network_loader import download_and_load_graph, get_network_list
from precompute_rankings import load_precomputed_rankings, get_standardized_ranked_nodes


# -----------------------------
# Configuration
# -----------------------------
MASTER_SEED = 42
GAMMA = 1
BETA_MULTIPLIER =2.5
SEED_RATIOS = [0.01, 0.05, 0.10]

FULL_METHOD = "HOSH"
VARIANTS = [
    "HOSH-NO",       # no overlap redundancy
    "HOSH-NE",       # no external-autonomy adjustment
    "HOSH-E",        # external-degree-only variant
    "HOSH-C",        # clique-membership/basic-dependence-only variant
    "HOSH-Lin",      # linear external-neighbor contribution
    "HOSH-Sqrt",     # alias of the Box-Cox concave variant in hosh_methods.py
    "HOSH-SumNorm",  # sum normalization instead of max normalization
]
METHODS = [FULL_METHOD] + VARIANTS

# Display names used in the manuscript. Internal code may still use HOSH
# if hosh_methods.py/precomputed rankings use the old method names.
DISPLAY_NAMES = {
    "HOSH": "MSH",
    "HOSH-NO": "MSH-NO",
    "HOSH-NE": "MSH-NE",
    "HOSH-E": "MSH-E",
    "HOSH-C": "MSH-C",
    "HOSH-Lin": "MSH-Lin",
    "HOSH-Sqrt": "MSH-AltConcave",
    "HOSH-SumNorm": "MSH-SumNorm",
}


def paired_wilcoxon_pvalue(x, y) -> float:
    """Two-sided paired Wilcoxon signed-rank test with safe handling of all-zero differences."""
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    mask = np.isfinite(x) & np.isfinite(y)
    x, y = x[mask], y[mask]
    if len(x) == 0:
        return np.nan
    diff = x - y
    if np.allclose(diff, 0):
        return 1.0
    try:
        return float(wilcoxon(x, y, zero_method="wilcox", alternative="two-sided").pvalue)
    except ValueError:
        return 1.0


def holm_adjust(p_values):
    """Holm-Bonferroni adjustment. NaN values are preserved."""
    p = np.asarray(p_values, dtype=float)
    adjusted = np.full_like(p, np.nan, dtype=float)
    valid = np.where(np.isfinite(p))[0]
    if len(valid) == 0:
        return adjusted.tolist()
    order = valid[np.argsort(p[valid])]
    m = len(order)
    prev = 0.0
    for rank, idx in enumerate(order):
        val = (m - rank) * p[idx]
        val = max(prev, min(val, 1.0))
        adjusted[idx] = val
        prev = val
    return adjusted.tolist()


def significance_stars(p: float) -> str:
    if not np.isfinite(p):
        return ""
    if p < 0.05:
        return "*"
    return ""


def format_improvement_with_sig(value: float, p: float) -> str:
    if not np.isfinite(value):
        return ""
    return f"{value:.2f}{significance_stars(p)}"


# -----------------------------
# Reproducible random seeds
# -----------------------------
def set_master_seed(seed: int = MASTER_SEED) -> None:
    random.seed(seed)
    np.random.seed(seed)


def stable_int_hash(*items, modulo=2**32 - 1) -> int:
    """Stable integer hash for deterministic seed derivation across Python processes."""
    text = "||".join(str(x) for x in items)
    digest = hashlib.blake2b(text.encode('utf-8'), digest_size=8).hexdigest()
    return int(digest, 16) % modulo


def shared_sir_seed(network_name: str, ratio_idx: int, block_idx: int, repeat_idx: int,
                    base_seed: int = MASTER_SEED) -> int:
    """
    Generate method-independent SIR random seed.
    The same seed is used for HOSH and all variants, enabling paired stochastic conditions.
    """
    return stable_int_hash(base_seed, network_name, ratio_idx, block_idx, repeat_idx)


# -----------------------------
# SIR model
# -----------------------------
def run_sir_simulation(graph, seeds: Iterable[int], beta: float, gamma: float,
                       rng_seed: int, max_steps: int = 1000) -> int:
    """Discrete-time synchronous SIR simulation under a local RNG."""
    rng = random.Random(rng_seed)

    infected = set(n for n in seeds if graph.has_node(n))
    recovered = set()

    if not infected:
        return 0

    for _ in range(max_steps):
        if not infected:
            break

        new_infected = set()
        new_recovered = set()

        for u in list(infected):
            for v in graph.neighbors(u):
                if v not in infected and v not in recovered:
                    if rng.random() < beta:
                        new_infected.add(v)

            if rng.random() < gamma:
                new_recovered.add(u)

        infected.update(new_infected)
        infected.difference_update(new_recovered)
        recovered.update(new_recovered)

    return len(recovered) + len(infected)


def epidemic_parameters(graph, gamma: float = GAMMA, multiplier: float = BETA_MULTIPLIER) -> Tuple[float, float]:
    degrees = [d for _, d in graph.degree()]
    k_mean = float(np.mean(degrees))
    k2_mean = float(np.mean([d ** 2 for d in degrees]))
    beta_th = gamma * (k_mean / (k2_mean - k_mean)) if (k2_mean - k_mean) > 0 else 0.0
    beta = multiplier * beta_th
    return beta, gamma


# -----------------------------
# Ranking and evaluation
# -----------------------------
def load_or_compute_scores(method: str, graph, network_name: str) -> Dict[int, float]:
    """Load precomputed ranking scores if available; otherwise compute them."""
    rankings = load_precomputed_rankings(network_name)
    if rankings:
        if method in rankings and rankings[method] is not None:
            return rankings[method]
        # Compatibility: older code may store HOSH-BoxCox instead of HOSH-Sqrt.
        if method == "HOSH-Sqrt" and "HOSH-BoxCox" in rankings and rankings["HOSH-BoxCox"] is not None:
            return rankings["HOSH-BoxCox"]

    return get_node_scores(method, graph)


def select_top_k(scores: Dict[int, float], k: int) -> List[int]:
    """Use the same deterministic ID-based tie-breaking as the revised main experiments."""
    ranked_nodes = get_standardized_ranked_nodes(scores, round_decimals=8)
    return ranked_nodes[:k]


def evaluate_seed_set_blocks(graph, seeds: List[int], beta: float, gamma: float,
                             network_name: str, ratio_idx: int,
                             blocks: int, repeats: int) -> np.ndarray:
    """
    Return block-level final infection scales in percent.
    Each block is the average of `repeats` SIR realizations.
    """
    n = graph.number_of_nodes()
    block_values = []

    for b in range(blocks):
        total = 0
        for r in range(repeats):
            seed = shared_sir_seed(network_name, ratio_idx, b, r)
            total += run_sir_simulation(graph, seeds, beta, gamma, rng_seed=seed)
        block_mean_pct = (total / repeats) / n * 100.0
        block_values.append(block_mean_pct)

    return np.asarray(block_values, dtype=float)


def mean_ci95(block_values: np.ndarray) -> Tuple[float, float]:
    mean_val = float(np.mean(block_values))
    if len(block_values) <= 1:
        return mean_val, 0.0
    std_val = float(np.std(block_values, ddof=1))
    ci95 = float(t.ppf(0.975, df=len(block_values) - 1) * std_val / np.sqrt(len(block_values)))
    return mean_val, ci95


def pct_improvement(full_value: float, variant_value: float) -> float:
    if variant_value == 0:
        return np.nan
    return (full_value - variant_value) / variant_value * 100.0


# -----------------------------
# Table formatting/export
# -----------------------------
def build_improvement_table(records: List[dict]) -> pd.DataFrame:
    """Build one table: Network, p, variants..., plus Average improvement row."""
    rows = []

    # Preserve standard network order and p order.
    networks = []
    for rec in records:
        if rec["Network"] not in networks:
            networks.append(rec["Network"])

    for net in networks:
        for ratio in SEED_RATIOS:
            sub = [r for r in records if r["Network"] == net and abs(r["p"] - ratio) < 1e-12]
            if not sub:
                continue

            row = {
                "Network": net,
                "p": ratio,
            }
            for variant in VARIANTS:
                match = [r for r in sub if r["Variant"] == variant]
                row[variant] = match[0]["Improvement_%"] if match else np.nan
            rows.append(row)

    df = pd.DataFrame(rows)

    # Add one final Average improvement row across all network-ratio pairs.
    avg_row = {"Network": "Average improvement", "p": ""}
    for variant in VARIANTS:
        avg_row[variant] = float(np.nanmean(df[variant].to_numpy(dtype=float)))
    df = pd.concat([df, pd.DataFrame([avg_row])], ignore_index=True)

    return df



def build_average_improvement_summary(records: List[dict], p_adjust: str = "none"):
    """
    Main manuscript summary:
    Rows are ablation variants, columns are p=1%, p=5%, p=10%.
    Each cell is the mean improvement rate over networks.
    Significance is computed by paired Wilcoxon tests over network-level paired means:
        {F_MSH(network), F_variant(network)}, n = number of networks.
    This avoids treating repeated SIR blocks as independent networks.
    """
    long_df = pd.DataFrame(records)
    rows = []

    for ratio in SEED_RATIOS:
        for variant in VARIANTS:
            sub = long_df[(long_df["Variant"] == variant) & (np.isclose(long_df["p"], ratio))].copy()
            if sub.empty:
                continue

            p_network = paired_wilcoxon_pvalue(
                sub["HOSH_F(%)"].to_numpy(dtype=float),
                sub["Variant_F(%)"].to_numpy(dtype=float),
            )

            rows.append({
                "Variant": variant,
                "Variant_Display": DISPLAY_NAMES.get(variant, variant),
                "p": ratio,
                "p_label": f"p={int(ratio * 100)}%",
                "Mean_Improvement(%)": float(np.nanmean(sub["Improvement_%"].to_numpy(dtype=float))),
                "Median_Improvement(%)": float(np.nanmedian(sub["Improvement_%"].to_numpy(dtype=float))),
                "Std_Improvement(%)": float(np.nanstd(sub["Improvement_%"].to_numpy(dtype=float), ddof=1)) if len(sub) > 1 else 0.0,
                "Networks": int(sub["Network"].nunique()),
                "Win_Networks": int(np.sum(sub["Improvement_%"].to_numpy(dtype=float) > 0)),
                "Network_Wilcoxon_p": p_network,
            })

    summary = pd.DataFrame(rows)

    if summary.empty:
        return pd.DataFrame(), pd.DataFrame(), long_df

    if p_adjust.lower() == "holm":
        summary["P_for_sig"] = holm_adjust(summary["Network_Wilcoxon_p"].tolist())
        summary["P_Adjustment"] = "Holm"
    else:
        summary["P_for_sig"] = summary["Network_Wilcoxon_p"]
        summary["P_Adjustment"] = "None"

    summary["Significance"] = summary["P_for_sig"].apply(significance_stars)
    summary["Cell"] = [
        format_improvement_with_sig(v, p)
        for v, p in zip(summary["Mean_Improvement(%)"], summary["P_for_sig"])
    ]

    manuscript = (
        summary.pivot(index="Variant_Display", columns="p_label", values="Cell")
               .reset_index()
               .rename(columns={"Variant_Display": "Variant"})
    )

    # Enforce row and column order.
    ordered_variants = [DISPLAY_NAMES[v] for v in VARIANTS]
    manuscript["Variant"] = pd.Categorical(manuscript["Variant"], categories=ordered_variants, ordered=True)
    manuscript = manuscript.sort_values("Variant").reset_index(drop=True)
    manuscript["Variant"] = manuscript["Variant"].astype(str)

    ordered_cols = ["Variant"] + [f"p={int(r * 100)}%" for r in SEED_RATIOS]
    manuscript = manuscript[[c for c in ordered_cols if c in manuscript.columns]]

    # Round numeric columns in long summary.
    for col in ["Mean_Improvement(%)", "Median_Improvement(%)", "Std_Improvement(%)", "Network_Wilcoxon_p", "P_for_sig"]:
        if col in summary.columns:
            summary[col] = summary[col].astype(float).round(6)

    return manuscript, summary, long_df


def export_excel(df: pd.DataFrame, output_path: str) -> None:
    """Export one styled Excel table."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Use openpyxl through pandas for convenient styling.
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Improvement", index=False)
        ws = writer.sheets["Improvement"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        avg_fill = PatternFill("solid", fgColor="FFF2CC")
        neg_fill = PatternFill("solid", fgColor="FCE4D6")
        pos_fill = PatternFill("solid", fgColor="E2F0D9")
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header style.
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Body style and percent formatting.
        max_row = ws.max_row
        max_col = ws.max_column
        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            is_avg = (row[0].value == "Average improvement")
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if is_avg:
                    cell.font = Font(bold=True)
                    cell.fill = avg_fill

            # p column display.
            p_cell = row[1]
            if isinstance(p_cell.value, (int, float)):
                p_cell.number_format = "0.0%"

            # Improvement columns.
            for cell in row[2:]:
                if isinstance(cell.value, (int, float)) and not np.isnan(cell.value):
                    cell.number_format = "0.00"
                    if not is_avg:
                        cell.fill = pos_fill if cell.value >= 0 else neg_fill

        # Merge Network cells for the three p rows of each network, like the example table.
        current_start = 2
        while current_start <= max_row - 1:  # exclude average row
            net_name = ws.cell(row=current_start, column=1).value
            current_end = current_start
            while current_end + 1 <= max_row - 1 and ws.cell(row=current_end + 1, column=1).value == net_name:
                current_end += 1
            if current_end > current_start:
                ws.merge_cells(start_row=current_start, start_column=1, end_row=current_end, end_column=1)
                ws.cell(row=current_start, column=1).alignment = Alignment(horizontal="center", vertical="center")
            current_start = current_end + 1

        # Column widths.
        widths = {
            1: 22,
            2: 10,
        }
        for col_idx in range(3, max_col + 1):
            widths[col_idx] = 15
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "C2"
        ws.sheet_view.showGridLines = False


def export_csv(df: pd.DataFrame, output_path: str) -> None:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df.to_csv(output_path, index=False, encoding="utf-8-sig")


# -----------------------------
# Main workflow
# -----------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Ablation improvement percentage table for HOSH variants.")
    parser.add_argument("--networks", nargs="+", default=None, help="Networks to run; default: all 9 networks.")
    parser.add_argument("--blocks", type=int, default=50, help="Number of SIR simulation blocks.")
    parser.add_argument("--repeats", type=int, default=20, help="SIR realizations per block.")
    parser.add_argument("--output-dir", default="results/exp_ablation_improvement_sig", help="Output directory.")
    parser.add_argument("--save-csv", action="store_true", help="Also save CSV copies of the main tables.")
    parser.add_argument("--p-adjust", choices=["none", "holm"], default="none", help="P-value adjustment used for significance stars in the manuscript summary table.")
    args = parser.parse_args()

    set_master_seed(MASTER_SEED)

    networks = args.networks if args.networks is not None else get_network_list()
    all_records = []
    detail_records = []

    print("=" * 72)
    print("Ablation improvement percentage table")
    print(f"Networks: {networks}")
    print(f"Seed ratios: {[f'{int(r*100)}%' for r in SEED_RATIOS]}")
    print(f"SIR protocol: {args.blocks} blocks × {args.repeats} repeats, master seed={MASTER_SEED}")
    print("=" * 72)

    for net_idx, net in enumerate(networks, 1):
        print(f"\n[{net_idx}/{len(networks)}] Network: {net}")
        graph = download_and_load_graph(net, verbose=False)
        if graph is None or graph.number_of_nodes() == 0:
            print(f"  [Skip] {net}: failed to load or empty graph.")
            continue

        n = graph.number_of_nodes()
        beta, gamma = epidemic_parameters(graph)
        print(f"  Nodes={n}, Edges={graph.number_of_edges()}, beta={beta:.6f}, gamma={gamma:.2f}")

        # Load/compute rankings once per network.
        rankings = {}
        for method in METHODS:
            print(f"  Loading/computing scores: {method}")
            rankings[method] = load_or_compute_scores(method, graph, net)

        for ratio_idx, ratio in enumerate(SEED_RATIOS):
            k = max(1, int(n * ratio))
            print(f"  Seed ratio p={int(ratio*100)}%, k={k}")

            method_final_mean = {}
            method_final_ci = {}
            method_final_blocks = {}

            for method in tqdm(METHODS, desc=f"    p={int(ratio*100)}%", leave=False):
                seeds = select_top_k(rankings[method], k)
                blocks = evaluate_seed_set_blocks(
                    graph, seeds, beta, gamma,
                    network_name=net,
                    ratio_idx=ratio_idx,
                    blocks=args.blocks,
                    repeats=args.repeats,
                )
                mean_val, ci_val = mean_ci95(blocks)
                method_final_mean[method] = mean_val
                method_final_ci[method] = ci_val
                method_final_blocks[method] = blocks

                detail_records.append({
                    "Network": net,
                    "p": ratio,
                    "Method": method,
                    "Final infection scale mean (%)": mean_val,
                    "95% CI": ci_val,
                    "k": k,
                })

            hosh_value = method_final_mean[FULL_METHOD]
            hosh_blocks = method_final_blocks[FULL_METHOD]

            print(f"    {DISPLAY_NAMES[FULL_METHOD]} F(tc)={hosh_value:.3f}%")
            for variant in VARIANTS:
                variant_value = method_final_mean[variant]
                variant_blocks = method_final_blocks[variant]
                imp = pct_improvement(hosh_value, variant_value)
                p_block = paired_wilcoxon_pvalue(hosh_blocks, variant_blocks)
                all_records.append({
                    "Network": net,
                    "p": ratio,
                    "p_label": f"p={int(ratio * 100)}%",
                    "Variant": variant,
                    "Variant_Display": DISPLAY_NAMES.get(variant, variant),
                    "HOSH_F(%)": hosh_value,
                    "Variant_F(%)": variant_value,
                    "Improvement_%": imp,
                    "Block_Wilcoxon_p": p_block,
                    "Block_Significance": significance_stars(p_block),
                    "Blocks": args.blocks,
                    "Repeats": args.repeats,
                })
                print(f"      vs {DISPLAY_NAMES.get(variant, variant):16s}: {imp:8.2f}%  p_block={p_block:.4g}")

    if not all_records:
        print("No records generated. Please check network loading and ranking files.")
        return

    # Old detailed one-table format: Network × p × Variant improvement.
    df_network_detail = build_improvement_table(all_records)

    # New manuscript table: average improvement over nine networks for each p and variant,
    # with significance computed over network-level paired means.
    manuscript_table, summary_long, improvement_long = build_average_improvement_summary(
        all_records, p_adjust=args.p_adjust
    )

    output_xlsx = os.path.join(args.output_dir, "ablation_average_improvement_significance.xlsx")
    os.makedirs(args.output_dir, exist_ok=True)

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        manuscript_table.to_excel(writer, sheet_name="Manuscript_Table", index=False)
        summary_long.to_excel(writer, sheet_name="Summary_Long", index=False)
        improvement_long.to_excel(writer, sheet_name="Network_Detail_With_BlockP", index=False)
        df_network_detail.to_excel(writer, sheet_name="Network_Improvement_Table", index=False)
        pd.DataFrame(detail_records).to_excel(writer, sheet_name="FinalSpread_Mean_CI", index=False)

        # Basic formatting.
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in writer.sheets.values():
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 32)
            ws.freeze_panes = "B2"
            ws.sheet_view.showGridLines = False

    if args.save_csv:
        manuscript_csv = os.path.join(args.output_dir, "ablation_average_improvement_significance.csv")
        summary_csv = os.path.join(args.output_dir, "ablation_average_improvement_summary_long.csv")
        manuscript_table.to_csv(manuscript_csv, index=False, encoding="utf-8-sig")
        summary_long.to_csv(summary_csv, index=False, encoding="utf-8-sig")
        print(f"CSV saved to: {manuscript_csv}")
        print(f"CSV saved to: {summary_csv}")

    print("\n" + "=" * 72)
    print("Done.")
    print("Main manuscript table:")
    print(manuscript_table.to_string(index=False))
    print(f"Excel saved to: {output_xlsx}")
    print("=" * 72)


if __name__ == "__main__":
    main()
